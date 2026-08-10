#!/usr/bin/env python3
"""
suivi2csv.py — convertit "Suivi Bib3 2025.xlsx" (MATRICE DE DISPONIBILITÉ
par SKU du BE) en suivi.csv exploitable par le pipeline.

SOURCE : assets/ref/Suivi Bib3 2025.xlsx, onglet unique porteur de
données "Feuil1" (1916 lignes x 59 colonnes ; "Feuil2"/"Feuil3" sont
vides, 1x1, ignorés). Deux lignes d'en-tête réelles :
  - ligne 1 : groupes de colonnes fusionnées (SAGE / GED / COMMUNICATION
    / PRODUCTION), purement décoratif, ignoré.
  - ligne 2 : intitulés de colonnes réels, utilisés ci-dessous.
Les données commencent en ligne 3 et se terminent en ligne 1915 (ligne
1916 = "TOTAL", un footer de synthèse, PAS une ligne de donnée — exclue).

PIÈGE #1 — LIGNES DE SECTION (PAS DES SKU) :
certaines lignes de la colonne A sont des titres de rubrique (ex.
"CORNICHES", "LES ROSACES", "MOULURES et ORNEMENTS") et non des
références produit. Elles sont repérables de façon FIABLE et
STRUCTURELLE (pas par heuristique de texte) : Excel les fusionne sur la
plage A:AL (`ws.merged_cells.ranges`, motif "A<n>:AL<n>") — confirmé
sur les 65 plages fusionnées du fichier réel, dont 61 suivent ce motif
exact. Ces lignes sont exclues du décompte SKU et marquées
`is_section_header=True` dans une colonne de contrôle séparée (pas
émise dans suivi.csv final, seulement utilisée pour filtrer).
Deux lignes supplémentaires ("DWG famille corniches", "DWG Collection
RIVOLI ", etc., lignes 5-6, 226, 348...) NE SONT PAS fusionnées mais ne
sont pas non plus de vrais SKU produit — elles portent une désignation
libre en colonne B ("famille complète sans angles"...) mais aucun code
SKU réel identifiable. RÈGLE (déterministe, pas de liste blanche/noire
de texte) : une ligne est retenue comme SKU UNIQUEMENT si elle n'est
pas une ligne de section fusionnée ET si sa valeur de colonne A,
normalisée, contient au moins un chiffre (`any(c.isdigit() for c in a)`)
— tous les vrais SKU observés (D100, M038, 1101, 50.050, TAL14, S601,
901.63...) contiennent un chiffre ; toutes les lignes "DWG famille..."
observées n'en contiennent aucun. Les rares lignes de titre non
fusionnées mais sans chiffre (aucune trouvée contenant un chiffre dans
ce fichier) seraient donc naturellement exclues par cette même règle,
jamais une coïncidence de correspondance de texte.

PIÈGE #2 — CELLULES BOOLÉENNES NON BOOLÉENNES :
les colonnes de disponibilité de format contiennent very majoritairement
1 (présent) ou vide/None (absent), mais PAS exclusivement — valeurs
réelles observées : 'OK', 'X', 'x', '0', 'incomplète', '1 à modifier ',
'à refaire suite modification', 'reste à faire passage LED', 'à
supprimer', ' ' (un espace). RÈGLE (imposée) : chaque colonne booléenne
`<nom>` produit DEUX colonnes en sortie :
  - `<nom>` (bool) : True si la valeur brute normalisée (str().strip())
    est dans TRUTHY_VALUES = {"1", "OK", "X"} (insensible à la casse) ;
    False sinon (y compris vide/None ET tout texte non reconnu comme
    "présent" — un commentaire libre n'est pas une confirmation de
    disponibilité, seulement un indice ambigu à visualiser dans _raw).
  - `<nom>_raw` (str) : la valeur brute telle quelle (repr sans les
    quotes Python, "" si vide/None) — AUCUNE cellule n'est booléanisée
    sans laisser de trace de sa valeur d'origine, cohérent avec le
    principe déjà appliqué à `unite`/`cote1_mm` dans catalogue2csv.py.

STATUT_GESTION (dérivé du LIBELLÉ SKU en colonne A, jamais du contenu
des colonnes de format) — motifs observés et vérifiés sur ce fichier
réel (voir recherches ci-dessous), dans cet ordre de priorité (le
premier motif qui matche l'emporte) :
  1. SUPPRIME            : "supprim"            (D531 SUPPRIME, D623
     "(supprimée en gestion)", D902 "supprimée de la gestion")
  2. ANNULEE              : "annul"              (1216 "Annulée")
  3. PAS_EN_GESTION       : "pas en gestion" / "pas  en gestion"
     (variantes d'espacement) (D119, D201, D202)
  4. SPECIAL_NON_REFERENCE: "spécial non référencé" / "special non
     référencé" (D559S, D602S, D603S, D659S, D660S, D661S, D662S,
     D719S... — 12 occurrences)
  5. NON_FABRIQUE         : "non fabriqué"       (50.CASS, 50.FES,
     50.MARR — non demandé explicitement par l'utilisateur mais motif
     réel présent dans le fichier, extrait par la même logique de
     détection de statut plutôt que silencieusement ignoré/mélangé à
     ACTIF)
  6. ACTIF (défaut)       : aucun des motifs ci-dessus.
Le SKU lui-même (colonne `sku`) conserve le libellé BRUT de la colonne A
tel quel (colonne `designation_sku_brute`) — la normalisation via
normalize_sku() est appliquée dans une colonne séparée `sku_normalise`,
jamais en écrasant le libellé source (ces libellés portent souvent le
motif de statut lui-même, ex. "D531 SUPPRIME", et une extraction propre
de la référence technique isolée n'est PAS demandée ici).

SORTIE : tools/dxf_pipeline/suivi.csv, une ligne par SKU (hors lignes de
section et hors ligne TOTAL), colonnes :
  sku, sku_normalise, designation, code_famille, section, statut_gestion,
  commentaire,
  jpg, jpg_raw, plan_pdf, plan_pdf_raw, dwg3d, dwg3d_raw, dxf, dxf_raw,
  dwg2d, dwg2d_raw, png3ds, png3ds_raw, "3dsmax", "3dsmax_raw",
  fbx_obj, fbx_obj_raw, rfa, rfa_raw, step, step_raw, skp, skp_raw,
  stl, stl_raw, ged2025, ged2025_raw, catalogue_nov2025,
  catalogue_nov2025_raw.

Plus un récapitulatif chiffré imprimé (nombre de SKU par format présent,
hors lignes de section/TOTAL).
"""
import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from dxf2profile import normalize_sku  # noqa: E402

HERE = Path(__file__).parent
DEFAULT_XLSX_PATH = HERE.parent.parent / "assets" / "ref" / "Suivi Bib3 2025.xlsx"
DEFAULT_OUTPUT_PATH = HERE / "suivi.csv"

SHEET_NAME = "Feuil1"
HEADER_ROW = 2
FIRST_DATA_ROW = 3

# colonne A -> col1, mapping 1-based Excel tel que vérifié sur le fichier
# réel (ligne d'en-tête 2 lue par openpyxl, voir docstring).
COL_DESIGNATION = 2       # B: "désignation de l'article"
COL_CODE_FAMILLE = 4      # D: "code famille "
COL_COMMENTAIRE_1 = 26    # Z: "COMMENTAIRES"
COL_COMMENTAIRE_2 = 38    # AL: "COMMENTAIRES" (deuxième bloc, PRODUCTION)

# nom_sortie -> colonne Excel (1-based), vérifié colonne par colonne sur
# la ligne d'en-tête réelle (ligne 2).
BOOL_COLUMNS = {
    "jpg": 5,               # JPG
    "plan_pdf": 6,           # plan PDF
    "dwg3d": 7,              # DWG3D
    "dxf": 8,                # dxf
    "dwg2d": 9,              # DWG 2D
    "png3ds": 10,            # rendu .PNG 3DS MAX
    "3dsmax": 11,            # 3DS MAX
    "fbx_obj": 12,           # fbx/obj/mtl
    "rfa": 13,               # RFA
    "step": 14,              # STEP
    "skp": 15,               # SKP
    "stl": 16,               # STL
    "ged2025": 22,           # GED 2025
    "catalogue_nov2025": 25,  # Intégration catalogue nov. 2025
}

TRUTHY_VALUES = {"1", "OK", "X"}

# Ordre de priorité : le premier motif qui matche l'emporte (cf.
# docstring — "supprim" doit primer sur un éventuel "non fabriqué" dans
# un même libellé, situation non observée mais réglée explicitement).
STATUT_PATTERNS = [
    ("SUPPRIME", re.compile(r"supprim", re.I)),
    ("ANNULEE", re.compile(r"annul", re.I)),
    ("PAS_EN_GESTION", re.compile(r"pas\s+en\s+gestion", re.I)),
    ("SPECIAL_NON_REFERENCE", re.compile(r"sp[ée]cial\s+non\s+r[ée]f[ée]renc", re.I)),
    ("NON_FABRIQUE", re.compile(r"non\s*fabriqu", re.I)),
]


def is_section_header_row(ws, row_num):
    """Une ligne est un titre de rubrique SI ET SEULEMENT SI elle fait
    partie d'une plage fusionnée couvrant A<row>:AL<row> — critère
    structurel Excel, jamais une devinette sur le texte. Vérifié: 61/65
    plages fusionnées du fichier réel suivent exactement ce motif."""
    target = f"A{row_num}:AL{row_num}"
    for rng in ws.merged_cells.ranges:
        if str(rng) == target:
            return True
    return False


def looks_like_real_sku(label):
    """Un vrai SKU produit contient toujours au moins un chiffre (vérifié
    sur D100, M038, 1101, 50.050, TAL14, S601, 901.63, SD-PRES1/2/3...).
    Les lignes de sous-titre non fusionnées ("DWG famille corniches",
    "DWG Collection RIVOLI ") n'en contiennent aucune — exclues par ce
    seul critère, sans liste blanche de texte."""
    if not label or not isinstance(label, str):
        return False
    return any(c.isdigit() for c in label)


def derive_statut_gestion(label):
    """Dérive statut_gestion du LIBELLÉ SKU (colonne A) uniquement,
    jamais du contenu des colonnes de disponibilité de format — cf.
    docstring module pour les motifs et leur ordre de priorité."""
    if not label:
        return "ACTIF"
    for statut, pattern in STATUT_PATTERNS:
        if pattern.search(label):
            return statut
    return "ACTIF"


def raw_str(value):
    """Représentation brute telle quelle, "" pour None/vide — jamais
    None dans le CSV (ambigu avec une vraie chaîne vide)."""
    if value is None:
        return ""
    return str(value)


def cell_is_truthy(value):
    """True si la valeur brute normalisée correspond à une confirmation
    explicite de disponibilité (TRUTHY_VALUES, insensible à la casse et
    aux espaces). Tout le reste (vide, texte libre ambigu, '0', 'x'
    minuscule inclus dans TRUTHY_VALUES via upper()) est False — un
    commentaire libre n'est PAS une confirmation, cf. docstring."""
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    return s.upper() in TRUTHY_VALUES


def build_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    current_section = ""
    rows = []
    nb_section_headers = 0
    nb_excluded_no_digit = 0
    nb_total_footer = 0

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value

        if label is not None and isinstance(label, str) and label.strip().upper() == "TOTAL":
            nb_total_footer += 1
            continue

        if is_section_header_row(ws, r):
            nb_section_headers += 1
            if label:
                current_section = label.strip()
            continue

        if not looks_like_real_sku(label):
            # ligne de sous-titre non fusionnée (ex. "DWG famille
            # corniches") ou ligne totalement vide — jamais un SKU.
            nb_excluded_no_digit += 1
            continue

        sku_brut = label.strip()
        row_out = {
            "sku": sku_brut,
            "sku_normalise": normalize_sku(sku_brut),
            "designation": raw_str(ws.cell(row=r, column=COL_DESIGNATION).value),
            "code_famille": raw_str(ws.cell(row=r, column=COL_CODE_FAMILLE).value),
            "section": current_section,
            "statut_gestion": derive_statut_gestion(sku_brut),
            "commentaire": " | ".join(
                filter(None, [
                    raw_str(ws.cell(row=r, column=COL_COMMENTAIRE_1).value),
                    raw_str(ws.cell(row=r, column=COL_COMMENTAIRE_2).value),
                ])
            ),
        }
        for name, col in BOOL_COLUMNS.items():
            raw_val = ws.cell(row=r, column=col).value
            row_out[name] = cell_is_truthy(raw_val)
            row_out[f"{name}_raw"] = raw_str(raw_val)

        rows.append(row_out)

    meta = {
        "nb_section_headers": nb_section_headers,
        "nb_excluded_no_digit": nb_excluded_no_digit,
        "nb_total_footer": nb_total_footer,
    }
    return rows, meta


def write_csv(rows, output_path):
    fieldnames = [
        "sku", "sku_normalise", "designation", "code_famille", "section",
        "statut_gestion", "commentaire",
    ]
    for name in BOOL_COLUMNS:
        fieldnames.append(name)
        fieldnames.append(f"{name}_raw")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX_PATH))
    ap.add_argument("--output", default=str(DEFAULT_OUTPUT_PATH))
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    output_path = Path(args.output)

    if not xlsx_path.exists():
        print(f"ERREUR: fichier introuvable: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    rows, meta = build_rows(xlsx_path)
    write_csv(rows, output_path)

    print(f"=== suivi.csv écrit ({len(rows)} SKU) -> {output_path} ===")
    print(f"  Lignes de section exclues (rubriques, fusion A:AL) : {meta['nb_section_headers']}")
    print(f"  Lignes exclues (sous-titre sans chiffre, pas un SKU) : {meta['nb_excluded_no_digit']}")
    print(f"  Ligne TOTAL (footer) exclue                          : {meta['nb_total_footer']}")

    print("\n--- Récapitulatif : nombre de SKU disposant de chaque format ---")
    for name in BOOL_COLUMNS:
        count = sum(1 for r in rows if r[name])
        print(f"  {name:20s}: {count}")

    statut_counts = {}
    for r in rows:
        statut_counts[r["statut_gestion"]] = statut_counts.get(r["statut_gestion"], 0) + 1
    print("\n--- Récapitulatif : statut_gestion ---")
    for statut, count in sorted(statut_counts.items(), key=lambda kv: -kv[1]):
        print(f"  {statut:25s}: {count}")


if __name__ == "__main__":
    main()
